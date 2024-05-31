import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import uuid

def generate_analysis_with_indicators(ticker_symbol, days=90, save_to_disk=True):
    # Calculate start and end dates
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=days)  # Considering data from the past 'days' days

    # Download data from Yahoo Finance
    data = yf.download(ticker_symbol, start=start_date, end=end_date)

    # Reset index to make Date a column instead of index
    data.reset_index(inplace=True)

    # Calculate EMAs
    data['EMA_5'] = data['Close'].ewm(span=5, adjust=False).mean()
    data['EMA_20'] = data['Close'].ewm(span=20, adjust=False).mean()
    data['EMA_12'] = data['Close'].ewm(span=12, adjust=False).mean()
    data['EMA_26'] = data['Close'].ewm(span=26, adjust=False).mean()

    # Calculate MACD and Signal line
    data['MACD'] = data['EMA_12'] - data['EMA_26']
    data['Signal'] = data['MACD'].ewm(span=9, adjust=False).mean()
    data['MACD_Histogram'] = data['MACD'] - data['Signal']

    # Calculate Pivot Points
    data['Pivot'] = (data['High'].shift(1) + data['Low'].shift(1) + data['Close'].shift(1)) / 3
    data['R1'] = 2 * data['Pivot'] - data['Low'].shift(1)
    data['S1'] = 2 * data['Pivot'] - data['High'].shift(1)
    data['R2'] = data['Pivot'] + (data['High'].shift(1) - data['Low'].shift(1))
    data['S2'] = data['Pivot'] - (data['High'].shift(1) - data['Low'].shift(1))

    # Calculate RSI
    delta = data['Close'].diff(1)
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)
    avg_gain = gain.rolling(window=14, min_periods=1).mean()
    avg_loss = loss.rolling(window=14, min_periods=1).mean()
    rs = avg_gain / avg_loss
    data['RSI'] = 100 - (100 / (1 + rs))

    # Calculate Stochastic Oscillator
    low_14 = data['Low'].rolling(window=14).min()
    high_14 = data['High'].rolling(window=14).max()
    data['%K'] = (data['Close'] - low_14) * 100 / (high_14 - low_14)
    data['%D'] = data['%K'].rolling(window=3).mean()

    # Determine individual buy signals
    data['MACD_Buy_Signal'] = data['MACD'] > data['Signal']
    data['Pivot_Buy_Signal'] = data['Close'] > data['Pivot']
    data['EMA_Buy_Signal'] = data['EMA_5'] > data['EMA_20']
    data['RSI_Buy_Signal'] = data['RSI'] < 70
    data['Stochastic_Buy_Signal'] = data['%K'] < 80

    # Combine all conditions to determine the overall buy signal
    data['Overall_Buy_Signal'] = (
        data['MACD_Buy_Signal'] &
        data['Pivot_Buy_Signal'] &
        data['EMA_Buy_Signal'] &
        data['RSI_Buy_Signal'] &
        data['Stochastic_Buy_Signal']
    )

    # Calculate the percentage of indicators giving a buy signal
    data['Buy_Signal_Percentage'] = (
        data[['MACD_Buy_Signal', 'Pivot_Buy_Signal', 'EMA_Buy_Signal', 'RSI_Buy_Signal', 'Stochastic_Buy_Signal']].sum(axis=1) / 5 * 100
    )

    if save_to_disk:
        # Generate a unique identifier (GUID)
        guid = uuid.uuid4().hex

        # Save the DataFrame to Excel with GUID appended to the filename
        output_file = f"Analysis_with_Indicators_{guid}.xlsx"
        data.to_excel(output_file, index=False)

        # Load the workbook and select the active worksheet
        wb = load_workbook(output_file)
        ws = wb.active

        # Find the correct column for 'Buy_Signal_Percentage'
        percentage_column = data.columns.get_loc('Buy_Signal_Percentage') + 1

        # Define the color scale rule for the Buy_Signal_Percentage column
        color_scale_rule = ColorScaleRule(
            start_type='percentile', start_value=0, start_color='FF0000',  # Red
            mid_type='percentile', mid_value=50, mid_color='FFFF00',      # Yellow
            end_type='percentile', end_value=100, end_color='00FF00'      # Green
        )

        # Apply the color scale rule to the Buy_Signal_Percentage column
        col_letter = ws.cell(row=1, column=percentage_column).column_letter
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{len(data)+1}', color_scale_rule)

        # Save the workbook
        wb.save(output_file)
        print(f"Analysis saved to {output_file}")
        return last_row
    else:
        # Return the last row of the DataFrame and print it
        last_row = data.iloc[-1]
        # print("Last row of the data:")
        # print(last_row)
        return last_row

# Example usage:
# Example list of ticker symbols
ticker_symbols = ["ADANIENT.BO", "HDFCBANK.NS", "TATAMOTORS.NS"]

# Create an empty DataFrame to store the results
results_df = pd.DataFrame(columns=['Ticker', 'Buy_Signal_Percentage'])

# Iterate over each ticker symbol
for symbol in ticker_symbols:
    print(f"Analyzing {symbol}...")
    last_row = generate_analysis_with_indicators(symbol, save_to_disk=False)
    
    # Append the results to the DataFrame
    results_df = results_df._append({'Date': datetime.now().date(), 'Ticker': symbol, 'Buy_Signal_Percentage': last_row['Buy_Signal_Percentage']}, ignore_index=True)

# Save the DataFrame to Excel
output_file = "Results_Analysis.xlsx"
results_df.to_excel(output_file, index=False)
print(f"Results saved to {output_file}")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import yfinance as yf
from datetime import datetime, timedelta

# Load your data
# data = pd.read_csv("C:\\data\\vishnu\\trading_with_gpt\\csv\\ADANIENT.BO.csv")
end_date = datetime.now().date()
start_date = end_date - timedelta(days=180)  # Considering data from the past 'days' days

ticker_symbol = 'TATAMOTORS.NS'
# Download data from Yahoo Finance
data = yf.download(ticker_symbol, start=start_date, end=end_date)

# Reset index to make Date a column instead of index
data.reset_index(inplace=True)

# Calculate EMAs
data['EMA_5'] = data['Close'].ewm(span=5, adjust=False).mean()
data['EMA_20'] = data['Close'].ewm(span=20, adjust=False).mean()
data['EMA_12'] = data['Close'].ewm(span=12, adjust=False).mean()
data['EMA_26'] = data['Close'].ewm(span=26, adjust=False).mean()

# Calculate MACD and Signal line
data['MACD'] = data['EMA_12'] - data['EMA_26']
data['Signal'] = data['MACD'].ewm(span=9, adjust=False).mean()
data['MACD_Histogram'] = data['MACD'] - data['Signal']

# Calculate Pivot Points
data['Pivot'] = (data['High'].shift(1) + data['Low'].shift(1) + data['Close'].shift(1)) / 3
data['R1'] = 2 * data['Pivot'] - data['Low'].shift(1)
data['S1'] = 2 * data['Pivot'] - data['High'].shift(1)
data['R2'] = data['Pivot'] + (data['High'].shift(1) - data['Low'].shift(1))
data['S2'] = data['Pivot'] - (data['High'].shift(1) - data['Low'].shift(1))

# Calculate RSI
delta = data['Close'].diff(1)
gain = delta.where(delta > 0, 0)
loss = -delta.where(delta < 0, 0)
avg_gain = gain.rolling(window=14, min_periods=1).mean()
avg_loss = loss.rolling(window=14, min_periods=1).mean()
rs = avg_gain / avg_loss
data['RSI'] = 100 - (100 / (1 + rs))

# Calculate Stochastic Oscillator
low_14 = data['Low'].rolling(window=14).min()
high_14 = data['High'].rolling(window=14).max()
data['%K'] = (data['Close'] - low_14) * 100 / (high_14 - low_14)
data['%D'] = data['%K'].rolling(window=3).mean()

# Determine individual buy signals
data['MACD_Buy_Signal'] = data['MACD'] > data['Signal']
data['Pivot_Buy_Signal'] = data['Close'] > data['Pivot']
data['EMA_Buy_Signal'] = data['EMA_5'] > data['EMA_20']
data['RSI_Buy_Signal'] = data['RSI'] < 70
data['Stochastic_Buy_Signal'] = data['%K'] < 80

# Combine all conditions to determine the overall buy signal
data['Overall_Buy_Signal'] = (
    data['MACD_Buy_Signal'] &
    data['Pivot_Buy_Signal'] &
    data['EMA_Buy_Signal'] &
    data['RSI_Buy_Signal'] &
    data['Stochastic_Buy_Signal']
)

# Calculate the percentage of indicators giving a buy signal
data['Buy_Signal_Percentage'] = (
    data[['MACD_Buy_Signal', 'Pivot_Buy_Signal', 'EMA_Buy_Signal', 'RSI_Buy_Signal', 'Stochastic_Buy_Signal']].sum(axis=1) / 5 * 100
)

# Determine potential buy points based on a threshold (e.g., 80%)
buy_threshold = 80
data['Potential_Buy'] = data['Buy_Signal_Percentage'] >= buy_threshold

# Simulate future price movements to check for a 4-5% gain
gain_target = 1.04  # 4% gain
data['Achieved_4_Percent_Gain'] = None

for i in range(len(data) - 1):
    if data.at[i, 'Potential_Buy']:
        # Assume its false
        data.at[i, 'Achieved_4_Percent_Gain'] = False
        for j in range(i + 1, len(data)):
            if data.at[j, 'Close'] >= data.at[i, 'Close'] * gain_target:
                data.at[i, 'Achieved_4_Percent_Gain'] = True
                break

# Save the DataFrame to Excel
output_file = "back_testing.xlsx"
data.to_excel(output_file, index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(output_file)
ws = wb.active

# Define the color scale rule for the Buy_Signal_Percentage column
color_scale_rule = ColorScaleRule(
    start_type='percentile', start_value=0, start_color='FF0000',  # Red
    mid_type='percentile', mid_value=50, mid_color='FFFF00',      # Yellow
    end_type='percentile', end_value=100, end_color='00FF00'      # Green
)

# Find the correct column for 'Buy_Signal_Percentage'
percentage_column = data.columns.get_loc('Buy_Signal_Percentage') + 1

# Apply the color scale rule to the Buy_Signal_Percentage column
col_letter = get_column_letter(percentage_column)
ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{len(data)+1}', color_scale_rule)

# Save the workbook
wb.save(output_file)
